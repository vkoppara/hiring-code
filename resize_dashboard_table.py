import argparse
from pathlib import Path

from openpyxl import load_workbook


def resize_table_from_start_cell(
    xlsx_path: str,
    sheet_name: str,
    start_cell: str = "A2",
    table_name: str | None = None,
) -> str:
    workbook = load_workbook(xlsx_path)
    sheet = workbook[sheet_name]

    if not sheet.tables:
        raise ValueError(f"No tables found in sheet '{sheet_name}'.")

    if table_name is None:
        table_name = next(iter(sheet.tables.keys()))

    if table_name not in sheet.tables:
        available = ", ".join(sheet.tables.keys())
        raise ValueError(
            f"Table '{table_name}' not found in sheet '{sheet_name}'. Available tables: {available}"
        )

    table = sheet.tables[table_name]

    start_col = "".join(ch for ch in start_cell if ch.isalpha()).upper()
    start_row_text = "".join(ch for ch in start_cell if ch.isdigit())
    if not start_col or not start_row_text:
        raise ValueError(f"Invalid start_cell '{start_cell}'. Example valid value: A2")
    start_row = int(start_row_text)

    current_end = table.ref.split(":")[1]
    end_col = "".join(ch for ch in current_end if ch.isalpha()).upper()

    last_row = sheet.max_row
    while last_row >= start_row:
        has_any_value = any(
            sheet.cell(row=last_row, column=col).value is not None
            for col in range(1, sheet.max_column + 1)
        )
        if has_any_value:
            break
        last_row -= 1

    if last_row < start_row:
        last_row = start_row

    table.ref = f"{start_col}{start_row}:{end_col}{last_row}"
    workbook.save(xlsx_path)
    return table.ref


def main() -> None:
    parser = argparse.ArgumentParser(description="Resize an Excel table after rows are added.")
    parser.add_argument("--file", required=True, help="Path to workbook (.xlsx)")
    parser.add_argument("--sheet", required=True, help="Worksheet name")
    parser.add_argument(
        "--start-cell",
        default="A2",
        help="Table header start cell (default: A2)",
    )
    parser.add_argument(
        "--table",
        default=None,
        help="Table name (optional). If omitted, first table in the sheet is used.",
    )
    args = parser.parse_args()

    workbook_path = Path(args.file)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    new_ref = resize_table_from_start_cell(
        xlsx_path=str(workbook_path),
        sheet_name=args.sheet,
        start_cell=args.start_cell,
        table_name=args.table,
    )
    print(f"Table resized to: {new_ref}")


if __name__ == "__main__":
    main()

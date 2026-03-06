import argparse
from pathlib import Path

from openpyxl import load_workbook


def _normalize_text(value) -> str:
    return str(value).strip().lower() if value is not None else ""


def _find_column_index(ws, header_row: int, expected_name: str) -> int:
    target = _normalize_text(expected_name)
    for col in range(1, ws.max_column + 1):
        if _normalize_text(ws.cell(row=header_row, column=col).value) == target:
            return col
    raise ValueError(f"Column '{expected_name}' not found in header row {header_row} of sheet '{ws.title}'.")


def _last_used_header_col(ws, header_row: int) -> int:
    last_col = 0
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value is not None:
            last_col = col
    return last_col if last_col > 0 else ws.max_column


def _last_data_row(ws, start_row: int, max_col: int) -> int:
    row = ws.max_row
    while row >= start_row:
        if any(ws.cell(row=row, column=col).value is not None for col in range(1, max_col + 1)):
            return row
        row -= 1
    return start_row - 1


def _ensure_target_headers(source_ws, target_ws, top_header_row: int, header_row: int, max_col: int) -> None:
    source_header = [_normalize_text(source_ws.cell(row=header_row, column=col).value) for col in range(1, max_col + 1)]
    target_header = [_normalize_text(target_ws.cell(row=header_row, column=col).value) for col in range(1, max_col + 1)]

    if any(target_header):
        if source_header != target_header:
            raise ValueError(
                f"Header mismatch between '{source_ws.title}' and '{target_ws.title}'. "
                "Please align columns before moving records."
            )
        return

    for col in range(1, max_col + 1):
        target_ws.cell(row=top_header_row, column=col).value = source_ws.cell(row=top_header_row, column=col).value
        target_ws.cell(row=header_row, column=col).value = source_ws.cell(row=header_row, column=col).value


def _resize_tables(ws) -> None:
    if not ws.tables:
        return
    for table_name in list(ws.tables.keys()):
        table = ws.tables[table_name]
        start_ref, end_ref = table.ref.split(":")
        start_col = "".join(ch for ch in start_ref if ch.isalpha())
        start_row = int("".join(ch for ch in start_ref if ch.isdigit()))
        end_col = "".join(ch for ch in end_ref if ch.isalpha())

        last_row = ws.max_row
        while last_row >= start_row:
            has_value = any(cell.value is not None for cell in ws[last_row])
            if has_value:
                break
            last_row -= 1
        if last_row < start_row:
            last_row = start_row
        table.ref = f"{start_col}{start_row}:{end_col}{last_row}"


def move_closed_filled_records(
    file_path: str,
    source_sheet: str = "Hiring Dashboard",
    target_sheet: str = "closed and filled",
    status_column: str = "Job Requisition Status",
    top_header_row: int = 1,
    header_row: int = 2,
) -> int:
    wb = load_workbook(file_path)
    if source_sheet not in wb.sheetnames:
        raise ValueError(f"Source sheet '{source_sheet}' not found.")

    source_ws = wb[source_sheet]
    target_ws = wb[target_sheet] if target_sheet in wb.sheetnames else wb.create_sheet(title=target_sheet)

    max_col = _last_used_header_col(source_ws, header_row)
    _ensure_target_headers(source_ws, target_ws, top_header_row, header_row, max_col)

    status_col_idx = _find_column_index(source_ws, header_row, status_column)
    data_start = header_row + 1

    rows_to_move = []
    source_last_row = _last_data_row(source_ws, data_start, max_col)
    for row in range(data_start, source_last_row + 1):
        status_value = _normalize_text(source_ws.cell(row=row, column=status_col_idx).value)
        if "closed" in status_value or "filled" in status_value:
            rows_to_move.append(row)

    if not rows_to_move:
        wb.save(file_path)
        return 0

    target_last = _last_data_row(target_ws, data_start, max_col)
    append_row = max(data_start, target_last + 1)

    for src_row in rows_to_move:
        for col in range(1, max_col + 1):
            target_ws.cell(row=append_row, column=col).value = source_ws.cell(row=src_row, column=col).value
        append_row += 1

    for src_row in reversed(rows_to_move):
        source_ws.delete_rows(src_row, 1)

    _resize_tables(source_ws)
    _resize_tables(target_ws)

    wb.save(file_path)
    return len(rows_to_move)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Move rows with closed/filled status from one worksheet to another."
    )
    parser.add_argument("--file", required=True, help="Path to workbook")
    parser.add_argument("--source", default="Hiring Dashboard", help="Source sheet name")
    parser.add_argument("--target", default="closed and filled", help="Target sheet name")
    parser.add_argument("--status-column", default="Job Requisition Status", help="Status column header")
    parser.add_argument("--top-header-row", type=int, default=1, help="Top group header row index")
    parser.add_argument("--header-row", type=int, default=2, help="Actual column header row index")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    moved = move_closed_filled_records(
        file_path=str(path),
        source_sheet=args.source,
        target_sheet=args.target,
        status_column=args.status_column,
        top_header_row=args.top_header_row,
        header_row=args.header_row,
    )
    print(f"Moved {moved} record(s) from '{args.source}' to '{args.target}'.")


if __name__ == "__main__":
    main()
